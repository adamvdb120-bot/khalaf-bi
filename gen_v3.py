import openpyxl
import random
from datetime import date, timedelta

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Data"

headers = [
    "Datum", "Dag", "Categorie", "Aantal_orders",
    "Omzet_excl_BTW", "BTW_9pct", "Omzet_incl_BTW",
    "Kosten_excl_BTW",
    "Personeelskosten", "Inkoopkosten", "Huurkosten",
    "Marketingkosten", "Overige_kosten",
    "Brutowinst_marge_pct"
]
ws.append(headers)

random.seed(42)
days_nl = ["Maandag","Dinsdag","Woensdag","Donderdag","Vrijdag","Zaterdag","Zondag"]
categories = ["Diner","Lunch","Afhaal","Bezorging"]
cat_factor = {"Diner": 1.0, "Lunch": 0.35, "Afhaal": 0.35, "Bezorging": 0.25}

start = date(2026, 1, 4)
for i in range(84):  # 12 weeks
    d = start + timedelta(days=i)
    dag = days_nl[d.weekday()]
    weekend = d.weekday() >= 5
    for cat in categories:
        base = random.uniform(15, 25) * (1.4 if weekend else 1.0) * cat_factor[cat]
        orders = max(1, int(random.gauss(base, base * 0.15)))
        omzet = round(orders * random.uniform(18, 28), 2)
        btw = round(omzet * 0.09, 2)
        incl = round(omzet + btw, 2)

        # Total kosten (~18-25% of omzet)
        kosten_pct = random.uniform(0.18, 0.25)
        kosten = round(omzet * kosten_pct, 2)

        # Split into categories (realistic horeca ratios)
        personeel  = round(kosten * random.uniform(0.42, 0.50), 2)
        inkoop     = round(kosten * random.uniform(0.28, 0.36), 2)
        huur       = round(kosten * random.uniform(0.10, 0.14), 2)
        marketing  = round(kosten * random.uniform(0.04, 0.07), 2)
        overig     = round(kosten - personeel - inkoop - huur - marketing, 2)

        marge = round(((omzet - kosten) / omzet) * 100, 1) if omzet > 0 else 0

        ws.append([
            d.strftime("%Y-%m-%d"), dag, cat, orders,
            omzet, btw, incl, kosten,
            personeel, inkoop, huur, marketing, overig,
            marge
        ])

output = r"C:\Users\adama\OneDrive - Windesheim Office365\Finance and Control jaar 4\Minor\oefening_klantdata_restaurant_v3.xlsx"
wb.save(output)
print(f"Saved: {output}")
print(f"Rows: {ws.max_row - 1}, Columns: {ws.max_column}")
